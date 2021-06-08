import streamlit as st
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import time
import SessionState
import helper
from random import randint

def main(session_state):

    st.markdown(""" <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    </style> """, unsafe_allow_html=True)

    hide_footer_style = """
    <style>
    .reportview-container .main footer {visibility: hidden;}    
    """
    st.markdown(hide_footer_style, unsafe_allow_html=True)

    primaryColor = st.get_option("theme.primaryColor")
    s = f"""
    <style>
    div.stButton > button:first-child {{ border: 5px solid {primaryColor}; border-radius:20px 20px 20px 20px; background-color: #0066cc; color:#ffffff}}
    <style>
    """
    st.markdown(s, unsafe_allow_html=True)

    st.title('Create Order')

    if session_state.successful_order == False:
        if session_state.prod_arr != [[], [], [], []]:
            st.subheader(' ')
            form_error = st.empty()
            st.echo()
            col1, col2 = st.beta_columns([50,50])

            if session_state.user_status == 'non_verify' or session_state.mobile == '':
                mob = col1.text_input("Enter Mobile number", key='mobile')
                if mob:
                    if len(mob) != 10 or mob.isnumeric()==False or len(mob) == 0 or mob is None:
                        form_error.error("Enter 10 digit valid Mobile number (India only)")
                        ok_for_order = False
                        st.stop()
                    else:
                        if session_state.mobile != str(mob):
                            otp = randint(10000, 99999)
                            helper.check_mob_otp(mob, otp)
                            sms_status = helper.send_otp(mob, otp)

                            if sms_status == 'true':
                                form_error.success("OTP sent to " + str(mob) + ", enter same in 'Enter OTP' box.")
                                session_state.mobile = str(mob)
                            else:
                                form_error.error("Mobile number is Invalid, Enter correct mobile number.")
                                st.stop()
                else:
                    st.stop()

                otp = col2.text_input("Enter OTP - sent on mobile number", key='otp')
                otp_status = ''

                if otp:
                    if len(otp) != 5 or otp.isnumeric() == False:
                        form_error.error("Enter 5 digit OTP sent on " + str(mob))
                        st.stop()
                    else:
                        otp_status = helper.check_otp_of_user(mob, otp)

                        if otp_status == False:
                            form_error.error("OTP mismatch!!!  Enter 5 digit OTP sent on " + str(mob) + '  \n NOTE: Incase if you haven\'t received OTP, Click Resend Button')

                            if session_state.resend > 0:
                                resend = st.button('Resend OTP')

                                if resend:
                                    session_state.resend = session_state.resend - 1
                                    form_error.success("OTP re-sent to " + str(mob) + ", enter same in 'Enter OTP' box.")
                                    otp = randint(10000, 99999)
                                    helper.check_mob_otp(mob, otp)
                            else:
                                st.subheader('Create offline order on message')
                                st.text_area('Send message (include Name, Address with pincode, email), and press \'CTRL+ENTER\'.', max_chars=150)
                                offline_order = st.button('Send message and create order offline')

                                if offline_order:
                                    session_state.successful_order = True
                                    session_state.prod_arr = [[], [], [], []]
                                    st.experimental_rerun()
                                    session_state.successful_order = False

                            st.stop()
                        else:
                            form_error.success("OTP matched! Mobile number verified successfully")
                            session_state.user_status = 'verified'
                else:
                    st.stop()
            else:
                col1.text_input("Enter Mobile number", key='mobile', value=session_state.mobile)
                col2.success("Mobile number already verified")

            st.markdown("""<hr>""", unsafe_allow_html=True)
            col5, col6 = st.beta_columns([50, 50])
            name = col5.text_input("Enter Name", key='name')
            pincode = col6.text_input("Enter pin code", key='pincode')
            email = col5.text_input("Enter email", key='email')
            address = col6.text_area("Enter Address", key='address')

            st.markdown("""<hr>""", unsafe_allow_html=True)
            total_header = st.empty()
            st.markdown("""<hr>""", unsafe_allow_html=True)

            clubbed_arr = []
            for i in range(len(session_state.prod_arr)):
                if len(session_state.prod_arr[i])>0:
                    for j in range(len(session_state.prod_arr[i])):
                        clubbed_arr.append(session_state.prod_arr[i][j])

            wb = load_workbook('productlog.xlsx')
            sheet = wb.active
            max_row = sheet.max_row
            total_final_actual_cost = 0
            for i in range(1,max_row):
                if sheet.cell(row=i+1, column=9).value in clubbed_arr:
                    col1, col2, col3, col4, col5 = st.beta_columns([10, 20, 10, 10, 10])
                    prod_code = str(sheet.cell(row=i + 1, column=9).value).strip()
                    col1.image('img_folder/' + prod_code + '.jpg')
                    col2.markdown('<sub><span style="color: #333333;">'+ str(sheet.cell(row=i+1, column=10).value).replace('<br>', '  \n')+'</span></sub>', unsafe_allow_html=True)
                    if str(sheet.cell(row=i+1, column=8).value).find('Free Size')== -1:
                        size = col3.selectbox("Select Size: ", options = ['Select'] + str(sheet.cell(row=i+1, column=8).value)[:-1].split(','), key = i)
                    else:
                        size = col3.selectbox("Select Size: ", options=str(sheet.cell(row=i + 1, column=8).value)[:-1].split(','), key = i)

                    qty = col4.selectbox("Select Qty: ", options = [1, 2, 3, 0], key = i)
                    col4.markdown('</span><span style="color: #0000ff;">select 0 if you dont want to order this product</span>', unsafe_allow_html=True)
                    col4.markdown("", unsafe_allow_html=True)
                    unit_cost = sheet.cell(row=i+1, column=4).value
                    final_actual_cost = qty*unit_cost
                    total_final_actual_cost = total_final_actual_cost + final_actual_cost
                    col5.subheader("Rs." + str(final_actual_cost))
                    st.markdown("""<hr>""", unsafe_allow_html=True)

            total_header.subheader("Total Cart value : Rs" + str(total_final_actual_cost) + '  \n  \n')
            #st.info(session_state.prod_arr[0])
            #st.info(session_state.prod_arr[1])
            #st.info(session_state.prod_arr[2])
            #st.info(session_state.prod_arr[3])
            #t.info(clubbed_arr)

            st.success('For Payment option, Only COD available!')

            form_error_end = st.empty()

            create_order = st.button('Create Order')
            ok_for_order = True

            if create_order:

                if len(name) < 6:
                    form_error.error("Enter full Name, so that it will be easy to locate your address")
                    ok_for_order = False
                    form_error_end.error("Enter full Name, so that it will be easy to locate your address")
                    st.stop()

                if len(pincode) != 6:
                    form_error.error("Enter 6 digit valid pincode, so that it will be easy to locate your address")
                    ok_for_order = False
                    form_error_end.error("Enter 6 digit valid pincode, so that it will be easy to locate your address")
                    st.stop()

                if len(email) == 0 or str(email).find('.')==-1 or str(email).find('@')==-1 or int(str(email).find('.'))<int(str(email).find('@')) or str(email).find('@')==0 or str(email).find('.')==len(email)-1:
                    form_error.error("Enter email in correct format")
                    ok_for_order = False
                    form_error_end.error("Enter email in correct format")
                    st.stop()

                if len(address) < 15:
                    form_error.error("Enter full Address, will help in successful order shipment")
                    ok_for_order = False
                    form_error_end.error("Enter full Address, will help in successful order shipment")
                    st.stop()

                if size == 'Select':
                    form_error.error("Select Size for all the Product(s)")
                    ok_for_order = False
                    form_error_end.error("Select Size for all the Product(s)")
                    st.stop()

                if ok_for_order == True:
                    session_state.successful_order = True
                    session_state.prod_arr = [[], [], [], []]
                    st.experimental_rerun()
                    session_state.successful_order = False

        else:
            st.warning('There is no product in cart.')
    else:
        st.success("Order created successfully, order id is 123456")
        st.warning('There is no product in cart.')
