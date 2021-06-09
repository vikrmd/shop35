import streamlit as st
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import time
import SessionState
import base64


def main(session_state):

    #st.sidebar.info('in start tshirt ' + str(session_state.prod_arr[0]))
    #st.sidebar.info('in start shirt ' + str(session_state.prod_arr[1]))
    #st.sidebar.info('in start sarees ' + str(session_state.prod_arr[2]))
    #st.sidebar.info('in start kurtis' + str(session_state.prod_arr[3]))

    # st.markdown(""" <style>
    # #MainMenu {visibility: hidden;}
    # footer {visibility: hidden;}
    # </style> """, unsafe_allow_html=True)
    #

    start_time = time.time()

    st.sidebar.header('Order Details')
    st.title('Baseoca')
    st.subheader('  ')


    wb = load_workbook('productlog.xlsx')
    sheet = wb.active
    max_row = sheet.max_row

    product_type = []

    for j in range(1, max_row):
        count = 0
        if sheet.cell(row=j+1, column=6).value not in product_type:
            product_type.append(sheet.cell(row=j+1, column=6).value)

    product_type.insert(0, 'Select')

    prod_typ = st.selectbox("Product Type: ", options = product_type)

    if prod_typ == 'tshirt':
        st.spinner()
        with st.spinner(text='In progress...loading ' + str(session_state.tshirt_count) + ' tshirts..'):
            time.sleep(2)
            ##st.success('Page loaded successfully')

    if prod_typ == 'shirts':
        st.spinner()
        with st.spinner(text='In progress...loading ' + str(session_state.shirts_count) + ' shirts..'):
            time.sleep(2)
            ##st.success('Page loaded successfully')

    if prod_typ == 'kurtis':
        st.spinner()
        with st.spinner(text='In progress...loading ' + str(session_state.shirts_count) + ' kurtis..'):
            time.sleep(2)
            ##st.success('Page loaded successfully')


    if prod_typ == 'sarees':
        st.spinner()
        with st.spinner(text='In progress...loading ' + str(session_state.shirts_count) + ' sarees..'):
            time.sleep(2)
            ##st.success('Page loaded successfully')


    curr_selected_prod = []
    # write the selected options
    if prod_typ == 'Select':
        st.warning("Select Product type from list to proceed...CART ITEMS (if present), will appear first")
        video_banner()
    else:
        st.info("Check sidebar (on left) to verify the products added to cart.")
        curr_selected_prod = load_products(prod_typ, session_state)
        #st.sidebar.info("Current selected prod " + str(curr_selected_prod))

    if prod_typ == 'tshirt':    session_state.prod_arr[0]=curr_selected_prod
    if prod_typ == 'shirts':    session_state.prod_arr[1]=curr_selected_prod
    if prod_typ == 'sarees':    session_state.prod_arr[2]=curr_selected_prod
    if prod_typ == 'kurtis':    session_state.prod_arr[3]=curr_selected_prod

    if session_state.prod_arr == [[],[],[],[]]:
        st.sidebar.info('Selected product(s) and cart value to be displayed here...')
    else:
        wb_info = load_workbook('productlog.xlsx')
        sheet_info = wb_info.active
        max_row = sheet_info.max_row
        session_state.successful_order = False

        for i in range(1, max_row):
            if sheet_info.cell(row=i + 1, column=9).value in session_state.prod_arr[0] or sheet_info.cell(row=i + 1, column=9).value in session_state.prod_arr[1] or sheet_info.cell(row=i + 1, column=9).value in session_state.prod_arr[2] or sheet_info.cell(row=i + 1, column=9).value in session_state.prod_arr[3]:
                st.sidebar.success(sheet_info.cell(row=i + 1, column=2).value + ' added to cart.  **Rs ' + str(sheet_info.cell(row=i + 1, column=4).value) + '**')

        wb_info.close()

    session_state.prev_prod_typ = prod_typ

    #st.sidebar.info('in end tshirt ' + str(session_state.prod_arr[0]))
    #st.sidebar.info('in end shirt ' + str(session_state.prod_arr[1]))
    #st.sidebar.info('in end sarees ' + str(session_state.prod_arr[2]))
    #t.sidebar.info('in end kurtis' + str(session_state.prod_arr[3]))

    end_time = time.time()
    #st.info('Response time : ' + str(end_time-start_time))
    #st.info(prod_typ)




#@st.cache(allow_output_mutation=True, suppress_st_warning=True)
def load_products(prod_typ, session_state):

    prod_loop = []

    wb = load_workbook('productlog.xlsx')
    sheet = wb.active
    max_row = sheet.max_row

    for i in range(1,max_row):

        if sheet.cell(row=i+1, column=6).value in prod_typ and (sheet.cell(row=i+1, column=9).value in session_state.prod_arr[0] or sheet.cell(row=i+1, column=9).value in session_state.prod_arr[1] or sheet.cell(row=i+1, column=9).value in session_state.prod_arr[2] or sheet.cell(row=i+1, column=9).value in session_state.prod_arr[3]):
            col1, mid, col2 = st.beta_columns([70,50,30])

            col1.subheader(sheet.cell(row=i+1, column=2).value)
            prod_code = str(sheet.cell(row=i+1, column=9).value).strip()
            col1.image('img_folder/'+ prod_code +'.jpg', use_column_width=80)
            mid.markdown('<sub><span style="color: #333333;">'+ str(sheet.cell(row=i+1, column=10).value).replace('<br>', '  \n')+'</span></sub>', unsafe_allow_html=True)
            col2.subheader("Rs." + str(sheet.cell(row=i+1, column=4).value))
            if sheet.cell(row=i + 1, column=9).value in session_state.prod_arr[0] or sheet.cell(row=i + 1, column=9).value in session_state.prod_arr[1] or sheet.cell(row=i + 1, column=9).value in session_state.prod_arr[2] or sheet.cell(row=i + 1, column=9).value in session_state.prod_arr[3]:
                add_to_cart = col2.checkbox("Add to Cart", key = i, value = True)
            else:
                add_to_cart = col2.checkbox("Add to Cart", key = i)

            if add_to_cart:
                prod_loop.append(sheet.cell(row=i+1, column=9).value)


            if str(sheet.cell(row=i+1, column=8).value).strip() != 'Free Size,':
                col2.markdown('<sub><span style="color: #333333;">Select size while creating an order.<br><br> Available size : ' + str(sheet.cell(row=i+1, column=8).value)[:-1]+'</span></sub>', unsafe_allow_html=True)
            else:
                col2.info("Size : " + str(sheet.cell(row=i + 1, column=8).value)[:-1])

            st.markdown("""<hr>""", unsafe_allow_html=True)

    for i in range(1, max_row):

        if sheet.cell(row=i+1, column=6).value in prod_typ and (sheet.cell(row=i+1, column=9).value not in session_state.prod_arr[0] and sheet.cell(row=i+1, column=9).value not in session_state.prod_arr[1] and sheet.cell(row=i+1, column=9).value not in session_state.prod_arr[2] and sheet.cell(row=i+1, column=9).value not in session_state.prod_arr[3]):
            col1, mid, col2 = st.beta_columns([60,60,30])

            prod_code = str(sheet.cell(row=i+1, column=9).value).strip()
            col1.image('img_folder/'+ prod_code +'.jpg')
            mid.markdown('<h4>'+sheet.cell(row=i+1, column=2).value + '</h4><br><sub><span style="color: #333333;">'+ str(sheet.cell(row=i+1, column=10).value)+'</span></sub>', unsafe_allow_html=True)
            col2.subheader("Rs." + str(sheet.cell(row=i+1, column=4).value))
            if sheet.cell(row=i + 1, column=9).value in session_state.prod_arr[0] or sheet.cell(row=i + 1, column=9).value in session_state.prod_arr[1] or sheet.cell(row=i + 1, column=9).value in session_state.prod_arr[2] or sheet.cell(row=i + 1, column=9).value in session_state.prod_arr[3]:
                add_to_cart = col2.checkbox("Add to Cart", key = i, value = True)
            else:
                add_to_cart = col2.checkbox("Add to Cart", key = i)

            if add_to_cart:
                prod_loop.append(sheet.cell(row=i+1, column=9).value)


            if str(sheet.cell(row=i+1, column=8).value).strip() != 'Free Size,':
                col2.markdown('<sub><span style="color: #1010F6;">Select size while creating an order.<br><br> Available size : ' + str(sheet.cell(row=i+1, column=8).value)[:-1]+'</span></sub>', unsafe_allow_html=True)
            else:
                col2.info("Size : " + str(sheet.cell(row=i + 1, column=8).value)[:-1])

            st.markdown("""<hr>""", unsafe_allow_html=True)

    return prod_loop


def video_banner():
    st.title('Demo')
    #video_file = open('how_to_order_product.mp4', 'rb')
    #video_bytes = video_file.read()
    #st.video(video_bytes)

    file_ = open("demo.gif", "rb")
    contents = file_.read()
    data_url = base64.b64encode(contents).decode("utf-8")
    file_.close()

    st.markdown(
        f'<img src="data:image/gif;base64,{data_url}" alt="cat gif" width="700">',
        unsafe_allow_html=True,
    )

if __name__ == "__main__":
    main()
