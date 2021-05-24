import streamlit as st
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import time
import SessionState

overall_cart = []
session_state = SessionState.get(prod_type=[], prev_mismatch=False, sarees_arr = [], kurtis_arr = [], shirts_arr = [], t_shirt_arr = [])


def main():
    st.set_page_config(page_title ='Shop35')

    st.markdown(""" <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    </style> """, unsafe_allow_html=True)

    st.sidebar.header('Order Details')
    st.title('SHOP35')
    st.subheader('  ')

    st.spinner()
    with st.spinner(text='In progress...'):
        time.sleep(2)
        ##st.success('Page loaded successfully')

    wb = load_workbook('productlog.xlsx')
    sheet = wb.active
    max_row = sheet.max_row

    product_type = []

    for j in range(1, max_row):
        if sheet.cell(row=j+1, column=6).value not in product_type:
            product_type.append(sheet.cell(row=j+1, column=6).value)

    product_type.insert(0, 'Select')

    prod_typ = st.selectbox("Product Type: ", options = product_type)

    # write the selected options
    if prod_typ == 'Select':
        st.warning("Kindly select Product type from list to proceed")
    else:
        st.info("Check sidebar (on left) to verify the products added to cart.")
        load_products(prod_typ)


    if len(session_state.sarees_arr) == 0 and len(session_state.kurtis_arr) == 0 and len(session_state.shirts_arr) == 0 and len(session_state.t_shirt_arr) == 0:
        st.sidebar.info('Selected product(s) and cart value to be displayed here...')
    else:
        display_order_info()
        #st.sidebar.info(session_state.sarees_arr)
        #st.sidebar.info(session_state.kurtis_arr)
        #st.sidebar.info(session_state.shirts_arr)
        #st.sidebar.info(session_state.t_shirt_arr)


def load_products(prod_typ):

    sarees_loop = []
    kurtis_loop = []
    shirts_loop = []
    t_shirt_loop = []

    wb = load_workbook('productlog.xlsx')
    sheet = wb.active
    max_row = sheet.max_row


    for i in range(1,max_row):

        if sheet.cell(row=i+1, column=6).value in prod_typ:
            col1, mid, col2 = st.beta_columns([80,30,50])

            col1.subheader(sheet.cell(row=i+1, column=2).value)
            prod_code = str(sheet.cell(row=i+1, column=9).value).strip()
            col1.image('img_folder/'+ prod_code +'.jpg', use_column_width=80)
            mid.text(sheet.cell(row=i+1, column=9).value)
            col2.subheader("Rs." + str(sheet.cell(row=i+1, column=4).value))

            if sheet.cell(row=i+1, column=9).value in session_state.sarees_arr or sheet.cell(row=i+1, column=9).value in session_state.kurtis_arr or sheet.cell(row=i+1, column=9).value in session_state.shirts_arr or sheet.cell(row=i+1, column=9).value in session_state.t_shirt_arr:
                add_to_cart = col2.checkbox("Add to Cart", key = i, value = True)
            else:
                add_to_cart = col2.checkbox("Add to Cart", key = i)

            st.markdown("""<hr>""", unsafe_allow_html=True)

            if str(sheet.cell(row=i+1, column=8).value).strip() != 'Free Size,':
                col2.info("Select size while creating an order.  \n  \n Available size : " + str(sheet.cell(row=i+1, column=8).value)[:-1])
            else:
                col2.info("Size : " + str(sheet.cell(row=i + 1, column=8).value)[:-1])

            if add_to_cart:
                if sheet.cell(row=i+1, column=6).value == 'sarees':
                    sarees_loop.append(sheet.cell(row=i+1, column=9).value)

                if sheet.cell(row=i+1, column=6).value == 'kurtis':
                    kurtis_loop.append(sheet.cell(row=i+1, column=9).value)

                if sheet.cell(row=i+1, column=6).value == 'shirts':
                    shirts_loop.append(sheet.cell(row=i+1, column=9).value)

                if sheet.cell(row=i+1, column=6).value == 'tshirt':
                    t_shirt_loop.append(sheet.cell(row=i+1, column=9).value)

    if session_state.prod_type != prod_typ:
        session_state.prev_mismatch = True
    else:
        session_state.prev_mismatch = False

    st.info(session_state.prod_type)
    st.info(prod_typ)

    session_state.prod_type = prod_typ
    st.warning(session_state.prev_mismatch)

    if 'sarees' in prod_typ and session_state.prev_mismatch != True:
        session_state.sarees_arr = sarees_loop

    if 'kurtis' in prod_typ and session_state.prev_mismatch != True:
        session_state.kurtis_arr = kurtis_loop

    if 'shirts' in prod_typ and session_state.prev_mismatch != True:
        session_state.shirts_arr = shirts_loop

    if 'tshirt' in prod_typ and session_state.prev_mismatch != True:
        session_state.t_shirt_arr = t_shirt_loop



def display_order_info():

    wb_info = load_workbook('productlog.xlsx')
    sheet_info = wb_info.active
    max_row = sheet_info.max_row


    if len(session_state.sarees_arr) != 0 or len(session_state.kurtis_arr) != 0 or len(session_state.shirts_arr) != 0 or len(session_state.t_shirt_arr) != 0:
        st.sidebar.info('Following are the products in your cart.  \n to create order, click \'Create Order\'')
        create_order = st.sidebar.button("Create Order")

    for i in range(1,max_row):
        if sheet_info.cell(row=i+1, column=9).value in session_state.shirts_arr:
            st.sidebar.success(sheet_info.cell(row=i+1, column=2).value + ' added to cart.  **Rs ' + str(sheet_info.cell(row=i+1, column=4).value) + '**')

        if sheet_info.cell(row=i+1, column=9).value in session_state.t_shirt_arr:
            st.sidebar.success(sheet_info.cell(row=i+1, column=2).value + ' added to cart.  **Rs ' + str(sheet_info.cell(row=i+1, column=4).value) + '**')

        if sheet_info.cell(row=i+1, column=9).value in session_state.sarees_arr:
            st.sidebar.success(sheet_info.cell(row=i+1, column=2).value + ' added to cart.  **Rs ' + str(sheet_info.cell(row=i+1, column=4).value) + '**')

        if sheet_info.cell(row=i+1, column=9).value in session_state.kurtis_arr:
            st.sidebar.success(sheet_info.cell(row=i+1, column=2).value + ' added to cart.  **Rs ' + str(sheet_info.cell(row=i+1, column=4).value) + '**')


if __name__ == "__main__":
    main()

