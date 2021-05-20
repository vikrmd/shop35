import streamlit as st
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import time
import SessionState

overall_cart = []
session_state = SessionState.get(prod_type=[], prev_mismatch=False, saree_arr = [], jewel_arr = [])


def main():

    st.markdown(""" <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    </style> """, unsafe_allow_html=True)

    st.sidebar.header('Order Details')

    st.title('SHOP35')
    st.subheader('  ')
    st.spinner()
    with st.spinner(text='In progress'):
        time.sleep(5)
        st.success('Page loaded successfully')

    wb = load_workbook('productlog.xlsx')
    sheet = wb.active
    max_row = sheet.max_row

    product_type = []

    for j in range(1, max_row):
        if sheet.cell(row=j+1, column=6).value not in product_type:
            product_type.append(sheet.cell(row=j+1, column=6).value)

    prod_typ = st.multiselect("Product Type: ", product_type)

    # write the selected options
    if len(prod_typ) == 0:
        st.warning("Kindly select atleast 1 product type from list to proceed")
    else:
        st.info("You selected " + str(len(prod_typ)) + ' Product type(s)')

    st.warning(prod_typ)

    load_products(prod_typ)


    if len(session_state.saree_arr) == 0 and len(session_state.jewel_arr) == 0:
        st.sidebar.info('Selected product(s) and cart value to be displayed here...')
    else:
        st.sidebar.info(session_state.saree_arr)
        st.sidebar.info(session_state.jewel_arr)


def load_products(prod_typ):

    saree_loop = []
    jewel_loop = []

    wb = load_workbook('productlog.xlsx')
    sheet = wb.active
    max_row = sheet.max_row

    for i in range(1,max_row):

        if sheet.cell(row=i+1, column=6).value in prod_typ:
            col1, mid, col2 = st.beta_columns([80,40,40])

            col1.subheader(sheet.cell(row=i+1, column=2).value)
            prod_code = str(sheet.cell(row=i+1, column=1).value).split('/')[-1].strip()
            col1.image('img_folder/'+ prod_code +'.png', use_column_width=80)
            mid.text(sheet.cell(row=i+1, column=8).value)
            col2.subheader("__Cost :__ Rs." + str(sheet.cell(row=i+1, column=4).value).replace('?',''))

            if sheet.cell(row=i+1, column=8).value in session_state.saree_arr or sheet.cell(row=i+1, column=8).value in session_state.jewel_arr:
                add_to_cart = col2.checkbox("Add to Cart", key = i, value = True)
            else:
                add_to_cart = col2.checkbox("Add to Cart", key = i)

            st.markdown("""<hr>""", unsafe_allow_html=True)

            if add_to_cart:
                if sheet.cell(row=i+1, column=6).value == 'sarees':
                    saree_loop.append(sheet.cell(row=i+1, column=8).value)

                if sheet.cell(row=i+1, column=6).value == 'sarees':
                    jewel_loop.append(sheet.cell(row=i+1, column=8).value)


    if session_state.prod_type != prod_typ:
        session_state.prev_mismatch = True
    else:
        session_state.prev_mismatch = False


    st.info(session_state.prod_type)
    st.info(prod_typ)

    session_state.prod_type = prod_typ
    st.warning(session_state.prev_mismatch)

    if 'sarees' in prod_typ and session_state.prev_mismatch != True:
        session_state.saree_arr = saree_loop


if __name__ == "__main__":
    main()

